import openpyxl

wk = openpyxl.load_workbook("C://Users//Patrick//Documents//robotframework_cours//Data_Pour_Test_TC_045.xlsx")

def fetch_number_of_rows(Sheetname):
    sh = wk[Sheetname]
    return sh.max_row

def fetch_cell_data(Sheetname, row, column):
    sh = wk[Sheetname]
    cell = sh.cell(row=int(row), column=int(column))
    return cell.value
